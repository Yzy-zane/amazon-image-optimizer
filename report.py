"""
Excel报告生成器
生成汽车装饰品图片质量分析报告，包含：
- 评分总表（带条件格式）
- 各维度详细评分
- 图表（评分分布、维度对比）
- 汇总分析Sheet
"""

import json
import os
import logging

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, PieChart, Reference, LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 样式定义
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FILL_2 = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="微软雅黑", size=10)
DATA_ALIGNMENT = Alignment(horizontal="center", vertical="center")
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def apply_cell_style(cell, font=None, fill=None, alignment=None, border=None):
    """应用单元格样式"""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def create_summary_sheet(wb, results):
    """创建评分总表"""
    ws = wb.active
    ws.title = "评分总表"

    # 列定义
    columns = [
        ("排名", 6),
        ("ASIN", 13),
        ("商品名称", 40),
        ("价格", 10),
        ("评分", 8),
        ("评论数", 10),
        ("图片数", 8),
        ("主图分辨率", 14),
        ("技术指标", 10),
        ("主图合规", 10),
        ("辅图丰富度", 12),
        ("设计质量", 10),
        ("营销效果", 10),
        ("总分", 10),
        ("百分制", 10),
    ]

    # 写入表头
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        apply_cell_style(cell, HEADER_FONT, HEADER_FILL, HEADER_ALIGNMENT, THIN_BORDER)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 写入数据
    for row_idx, r in enumerate(results, 2):
        tech = r.get("technical", {})
        mc = r.get("main_compliance", {})
        sr = r.get("sub_richness", {})
        dq = r.get("design_quality", {})
        mk = r.get("marketing", {})

        data = [
            row_idx - 1,
            r.get("asin", ""),
            r.get("title", "N/A")[:60],
            r.get("price", "N/A"),
            r.get("rating", "N/A"),
            r.get("reviews", "0"),
            r.get("image_count", 0),
            tech.get("resolution", "N/A"),
            tech.get("subtotal", 0),
            mc.get("subtotal", 0),
            sr.get("subtotal", 0),
            dq.get("subtotal", 0),
            mk.get("subtotal", 0),
            r.get("total_score", 0),
            r.get("normalized_score", 0),
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 3:
                apply_cell_style(cell, DATA_FONT, alignment=LEFT_ALIGNMENT, border=THIN_BORDER)
            else:
                apply_cell_style(cell, DATA_FONT, alignment=DATA_ALIGNMENT, border=THIN_BORDER)

    # 条件格式 - 百分制评分列(O列)
    last_row = len(results) + 1
    if last_row > 1:
        score_range = f"O2:O{last_row}"
        ws.conditional_formatting.add(
            score_range,
            ColorScaleRule(
                start_type="min", start_color="FFC7CE",
                mid_type="percentile", mid_value=50, mid_color="FFEB9C",
                end_type="max", end_color="C6EFCE"
            )
        )

        # 总分列(N列)也添加条件格式
        total_range = f"N2:N{last_row}"
        ws.conditional_formatting.add(
            total_range,
            ColorScaleRule(
                start_type="min", start_color="FFC7CE",
                mid_type="percentile", mid_value=50, mid_color="FFEB9C",
                end_type="max", end_color="C6EFCE"
            )
        )

    return ws


def create_detail_sheet(wb, results):
    """创建各维度详细评分Sheet"""
    ws = wb.create_sheet("详细评分")

    # 列定义 - 展开所有子维度
    columns = [
        ("ASIN", 13),
        # 技术指标
        ("分辨率评分\n(0-10)", 12),
        ("文件大小评分\n(0-5)", 12),
        ("图片数量评分\n(0-10)", 12),
        ("格式质量\n(0-5)", 10),
        # 主图合规
        ("白底评分\n(0-10)", 10),
        ("产品占比\n(0-10)", 10),
        ("文字检测\n(0-5)", 10),
        ("阴影检测\n(0-5)", 10),
        # 辅图丰富度
        ("多角度\n(0-10)", 10),
        ("场景图\n(0-5)", 10),
        ("信息图\n(0-5)", 10),
        ("细节图\n(0-5)", 10),
        # 设计质量
        ("光线均匀\n(1-5)", 10),
        ("构图平衡\n(1-5)", 10),
        ("色彩表现\n(1-5)", 10),
        ("清晰度\n(1-5)", 10),
        # 营销效果
        ("卖点突出\n(1-5)", 10),
        ("视觉吸引\n(1-5)", 10),
        ("场景代入\n(1-5)", 10),
        ("信息完整\n(1-5)", 10),
    ]

    # 二级表头 - 维度分组
    group_headers = [
        (1, 1, ""),
        (2, 5, "技术指标"),
        (6, 9, "主图合规性"),
        (10, 13, "辅图丰富度"),
        (14, 17, "设计质量"),
        (18, 21, "营销效果"),
    ]

    # 写入分组表头（第1行）
    for start, end, name in group_headers:
        if name:
            cell = ws.cell(row=1, column=start, value=name)
            apply_cell_style(cell, HEADER_FONT, HEADER_FILL, HEADER_ALIGNMENT, THIN_BORDER)
            if start != end:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
                for c in range(start + 1, end + 1):
                    apply_cell_style(ws.cell(row=1, column=c), HEADER_FONT, HEADER_FILL, HEADER_ALIGNMENT, THIN_BORDER)
        else:
            cell = ws.cell(row=1, column=start, value="")
            apply_cell_style(cell, HEADER_FONT, HEADER_FILL, HEADER_ALIGNMENT, THIN_BORDER)

    # 写入子维度表头（第2行）
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        apply_cell_style(cell, HEADER_FONT, HEADER_FILL_2, HEADER_ALIGNMENT, THIN_BORDER)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = "A3"
    ws.row_dimensions[2].height = 35

    # 写入数据
    for row_idx, r in enumerate(results, 3):
        tech = r.get("technical", {})
        mc = r.get("main_compliance", {})
        sr = r.get("sub_richness", {})
        dq = r.get("design_quality", {})
        mk = r.get("marketing", {})

        data = [
            r.get("asin", ""),
            tech.get("resolution_score", 0),
            tech.get("file_size_score", 0),
            tech.get("image_count_score", 0),
            tech.get("format_score", 0),
            mc.get("background_score", 0),
            mc.get("product_ratio_score", 0),
            mc.get("text_logo_score", 0),
            mc.get("shadow_score", 0),
            sr.get("multi_angle_score", 0),
            sr.get("scene_score", 0),
            sr.get("infographic_score", 0),
            sr.get("detail_score", 0),
            dq.get("lighting_score", 0),
            dq.get("composition_score", 0),
            dq.get("color_score", 0),
            dq.get("sharpness_score", 0),
            mk.get("selling_point_score", 0),
            mk.get("visual_appeal_score", 0),
            mk.get("scene_immersion_score", 0),
            mk.get("info_completeness_score", 0),
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_cell_style(cell, DATA_FONT, alignment=DATA_ALIGNMENT, border=THIN_BORDER)

    # 对评分列添加条件格式
    last_row = len(results) + 2
    if last_row > 2:
        for col in range(2, 22):
            col_letter = get_column_letter(col)
            range_str = f"{col_letter}3:{col_letter}{last_row}"
            ws.conditional_formatting.add(
                range_str,
                ColorScaleRule(
                    start_type="min", start_color="FFC7CE",
                    mid_type="percentile", mid_value=50, mid_color="FFEB9C",
                    end_type="max", end_color="C6EFCE"
                )
            )

    return ws


def create_charts_sheet(wb, results):
    """创建图表Sheet"""
    ws = wb.create_sheet("图表分析")

    if not results:
        ws.cell(row=1, column=1, value="没有数据可供分析")
        return ws

    # ---- 数据准备 ----
    # 1. 评分分布数据
    scores = [r.get("normalized_score", 0) for r in results]
    bins = ["0-20", "20-40", "40-60", "60-80", "80-100"]
    bin_counts = [0, 0, 0, 0, 0]
    for s in scores:
        if s < 20:
            bin_counts[0] += 1
        elif s < 40:
            bin_counts[1] += 1
        elif s < 60:
            bin_counts[2] += 1
        elif s < 80:
            bin_counts[3] += 1
        else:
            bin_counts[4] += 1

    # 写入评分分布数据
    ws.cell(row=1, column=1, value="评分区间")
    ws.cell(row=1, column=2, value="商品数量")
    for i, (b, c) in enumerate(zip(bins, bin_counts)):
        ws.cell(row=i + 2, column=1, value=b)
        ws.cell(row=i + 2, column=2, value=c)

    # 创建评分分布柱状图
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "商品图片质量评分分布"
    chart1.x_axis.title = "分数区间"
    chart1.y_axis.title = "商品数量"
    chart1.style = 10

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=6)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=6)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    chart1.shape = 4
    chart1.width = 18
    chart1.height = 12

    ws.add_chart(chart1, "D1")

    # 2. 各维度平均分对比
    dim_names = ["技术指标", "主图合规", "辅图丰富度", "设计质量", "营销效果"]
    dim_keys = ["technical", "main_compliance", "sub_richness", "design_quality", "marketing"]
    dim_max = [30, 30, 25, 20, 20]  # 各维度满分

    avg_scores = []
    for key in dim_keys:
        dim_totals = [r.get(key, {}).get("subtotal", 0) for r in results]
        avg_scores.append(round(sum(dim_totals) / max(len(dim_totals), 1), 1))

    # 写入维度对比数据
    start_row = 9
    ws.cell(row=start_row, column=1, value="评分维度")
    ws.cell(row=start_row, column=2, value="平均分")
    ws.cell(row=start_row, column=3, value="满分")
    for i, (name, avg, mx) in enumerate(zip(dim_names, avg_scores, dim_max)):
        ws.cell(row=start_row + 1 + i, column=1, value=name)
        ws.cell(row=start_row + 1 + i, column=2, value=avg)
        ws.cell(row=start_row + 1 + i, column=3, value=mx)

    # 创建维度对比柱状图
    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "各评分维度平均分 vs 满分"
    chart2.x_axis.title = "维度"
    chart2.y_axis.title = "分数"
    chart2.style = 10

    data_ref2 = Reference(ws, min_col=2, max_col=3, min_row=start_row, max_row=start_row + 5)
    cats_ref2 = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + 5)
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats_ref2)
    chart2.width = 18
    chart2.height = 12

    ws.add_chart(chart2, "D18")

    # 3. Top10 vs Bottom10 对比
    top10 = results[:10] if len(results) >= 10 else results
    bottom10 = results[-10:] if len(results) >= 10 else results

    start_row2 = 16
    ws.cell(row=start_row2, column=1, value="对比项目")
    ws.cell(row=start_row2, column=2, value="Top 10平均")
    ws.cell(row=start_row2, column=3, value="Bottom 10平均")

    for i, (name, key) in enumerate(zip(dim_names, dim_keys)):
        ws.cell(row=start_row2 + 1 + i, column=1, value=name)
        top_avg = sum(r.get(key, {}).get("subtotal", 0) for r in top10) / max(len(top10), 1)
        bot_avg = sum(r.get(key, {}).get("subtotal", 0) for r in bottom10) / max(len(bottom10), 1)
        ws.cell(row=start_row2 + 1 + i, column=2, value=round(top_avg, 1))
        ws.cell(row=start_row2 + 1 + i, column=3, value=round(bot_avg, 1))

    return ws


def create_analysis_sheet(wb, results):
    """创建汇总分析Sheet"""
    ws = wb.create_sheet("汇总分析")

    if not results:
        ws.cell(row=1, column=1, value="没有数据可供分析")
        return ws

    # 标题样式
    title_font = Font(name="微软雅黑", size=14, bold=True, color="2F5496")
    subtitle_font = Font(name="微软雅黑", size=12, bold=True, color="4472C4")
    content_font = Font(name="微软雅黑", size=10)

    row = 1
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    # 标题
    ws.cell(row=row, column=1, value="Amazon汽车装饰品图片质量分析报告").font = title_font
    ws.merge_cells("A1:B1")
    row += 2

    # 基本统计
    ws.cell(row=row, column=1, value="一、基本统计").font = subtitle_font
    row += 1

    total = len(results)
    scores = [r.get("normalized_score", 0) for r in results]
    avg_score = sum(scores) / max(total, 1)
    max_score = max(scores) if scores else 0
    min_score = min(scores) if scores else 0

    stats = [
        ("分析商品总数", f"{total} 个"),
        ("平均图片质量评分", f"{avg_score:.1f} / 100"),
        ("最高评分", f"{max_score:.1f} / 100"),
        ("最低评分", f"{min_score:.1f} / 100"),
        ("评分中位数", f"{sorted(scores)[len(scores)//2]:.1f} / 100" if scores else "N/A"),
    ]

    for label, value in stats:
        ws.cell(row=row, column=1, value=label).font = content_font
        ws.cell(row=row, column=2, value=value).font = content_font
        row += 1

    row += 1

    # 图片数量统计
    ws.cell(row=row, column=1, value="二、图片数量统计").font = subtitle_font
    row += 1

    img_counts = [r.get("image_count", 0) for r in results]
    avg_imgs = sum(img_counts) / max(total, 1)
    has_7plus = sum(1 for c in img_counts if c >= 7)
    has_5plus = sum(1 for c in img_counts if c >= 5)

    img_stats = [
        ("平均图片数", f"{avg_imgs:.1f} 张"),
        ("7张以上图片的商品", f"{has_7plus} 个 ({has_7plus/max(total,1)*100:.0f}%)"),
        ("5张以上图片的商品", f"{has_5plus} 个 ({has_5plus/max(total,1)*100:.0f}%)"),
    ]

    for label, value in img_stats:
        ws.cell(row=row, column=1, value=label).font = content_font
        ws.cell(row=row, column=2, value=value).font = content_font
        row += 1

    row += 1

    # 主图合规性统计
    ws.cell(row=row, column=1, value="三、主图合规性分析").font = subtitle_font
    row += 1

    bg_scores = [r.get("main_compliance", {}).get("background_score", 0) for r in results]
    high_bg = sum(1 for s in bg_scores if s >= 8)

    compliance_stats = [
        ("白底合规率（≥8分）", f"{high_bg} 个 ({high_bg/max(total,1)*100:.0f}%)"),
        ("平均白底评分", f"{sum(bg_scores)/max(total,1):.1f} / 10"),
        ("平均产品占比评分",
         f"{sum(r.get('main_compliance', {}).get('product_ratio_score', 0) for r in results)/max(total,1):.1f} / 10"),
    ]

    for label, value in compliance_stats:
        ws.cell(row=row, column=1, value=label).font = content_font
        ws.cell(row=row, column=2, value=value).font = content_font
        row += 1

    row += 1

    # 热销商品特征
    ws.cell(row=row, column=1, value="四、热销商品图片共同特征").font = subtitle_font
    row += 1

    # 按评论数排序取Top20
    by_reviews = sorted(results, key=lambda x: int(str(x.get("reviews", "0")).replace(",", "") or "0"), reverse=True)
    top_sellers = by_reviews[:20] if len(by_reviews) >= 20 else by_reviews

    if top_sellers:
        ts_avg_score = sum(r.get("normalized_score", 0) for r in top_sellers) / len(top_sellers)
        ts_avg_imgs = sum(r.get("image_count", 0) for r in top_sellers) / len(top_sellers)
        ts_scene = sum(r.get("sub_richness", {}).get("scene_count", 0) for r in top_sellers) / len(top_sellers)
        ts_info = sum(r.get("sub_richness", {}).get("infographic_count", 0) for r in top_sellers) / len(top_sellers)

        features = [
            ("热销商品(Top20)平均图片评分", f"{ts_avg_score:.1f} / 100"),
            ("热销商品平均图片数", f"{ts_avg_imgs:.1f} 张"),
            ("热销商品平均场景图数", f"{ts_scene:.1f} 张"),
            ("热销商品平均信息图数", f"{ts_info:.1f} 张"),
        ]

        for label, value in features:
            ws.cell(row=row, column=1, value=label).font = content_font
            ws.cell(row=row, column=2, value=value).font = content_font
            row += 1

    row += 1

    # 建议
    ws.cell(row=row, column=1, value="五、优化建议").font = subtitle_font
    row += 1

    suggestions = [
        "1. 确保主图使用纯白背景(RGB: 255,255,255)，产品占比60-85%",
        "2. 提供至少7张图片（1张主图 + 6张辅图），覆盖多个角度",
        "3. 辅图应包含：场景图、信息图/卖点图、细节特写图",
        "4. 图片分辨率建议≥2000x2000像素，确保在移动端缩放后仍然清晰",
        "5. 保持光线均匀，构图居中，色彩自然真实",
        "6. 信息图应突出核心卖点，使用对比色和清晰文字",
        "7. 场景图应展示产品在实际使用环境中的效果，增强代入感",
    ]

    for suggestion in suggestions:
        ws.cell(row=row, column=1, value=suggestion).font = content_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1

    return ws


def generate_report(analysis_file="analysis_results.json", output_file="汽车装饰品图片质量分析.xlsx"):
    """生成完整的Excel分析报告"""
    analysis_path = os.path.join(BASE_DIR, analysis_file)
    output_path = os.path.join(BASE_DIR, output_file)

    if not os.path.exists(analysis_path):
        logger.error(f"分析结果文件不存在: {analysis_path}")
        return None

    with open(analysis_path, "r", encoding="utf-8") as f:
        results = json.load(f)

    logger.info(f"加载 {len(results)} 条分析结果，开始生成报告...")

    wb = Workbook()

    # 创建各Sheet
    create_summary_sheet(wb, results)
    create_detail_sheet(wb, results)
    create_charts_sheet(wb, results)
    create_analysis_sheet(wb, results)

    # 保存
    wb.save(output_path)
    logger.info(f"报告已生成: {output_path}")
    return output_path


if __name__ == "__main__":
    generate_report()
